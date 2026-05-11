#!/usr/bin/env python3
"""
ESG Framework Scraper + Semantic Validation Agent
==================================================
Architecture:
  Question → Classification → Section Routing → Source Selection
  → Extraction Strategy → Scrape (gemini-2.5-flash, 3 workers)
  → Validate (gemini-2.5-pro) → Score → Final E/S/G + Overall

Features:
  - Question Resource Map: every question knows WHERE to search,
    WHAT keywords to use, WHAT table sections to look in,
    and WHAT answer mode to use (table_lookup | semantic_extraction | boolean_detection)
  - Semantic constraints: qualitative min_words, numeric must_contain_number
  - Fallback leakage detection across consecutive questions
  - 3 parallel workers for scraping
  - Final scores: E score, S score, G score, Overall ESG score

Usage:
  .venv/bin/python scripts/esg_framework_scraper.py \\
      --company "Tata Steel" --output reports/tata_steel_esg.json

  .venv/bin/python scripts/esg_framework_scraper.py \\
      --company "Infosys" "Reliance Industries" --output reports/multi.json

  .venv/bin/python scripts/esg_framework_scraper.py \\
      --company "Tata Steel" --no-validate --output reports/out.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

def find_project_root():
    current = Path(__file__).resolve().parent
    # Check current and up to 2 levels up for 'backend'
    for _ in range(3):
        if (current / "backend").exists():
            return current
        current = current.parent
    # Fallback to original parents[1] logic if not found
    return Path(__file__).resolve().parents[1]

PROJECT_ROOT = find_project_root()
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(str(PROJECT_ROOT / ".env"), override=False)
load_dotenv(str(PROJECT_ROOT / "backend" / ".env"), override=False)


# =============================================================================
# CONSTANTS & SCORING WEIGHTS
# =============================================================================

PILLAR_WEIGHTS: Dict[str, Dict[str, Any]] = {
    "Environment": {"pillar_weight": 0.40, "framework_max": 130},
    "Social":      {"pillar_weight": 0.35, "framework_max": 148},
    "Governance":  {"pillar_weight": 0.25, "framework_max": 79},
}

ESG_CATEGORY_LABELS = {
    "Environment": "Environmental",
    "Social":      "Social",
    "Governance":  "Governance",
}

ESG_PROFILES = [
    (0,  25,  "Foundational"),
    (25, 50,  "Developing"),
    (50, 75,  "Established"),
    (75, 101, "Advanced"),
]

# Attribute → question_type classification
QUALITATIVE_ATTRS = {
    "policy", "actions", "initiatives", "strategy",
    "governance", "commitment", "strategic",
}
QUANTITATIVE_ATTRS = {
    "performance", "consumption", "withdrawal", "emissions",
    "energy", "waste", "intensity", "perfomance",
}

# Semantic constraints per question_type
SEMANTIC_CONSTRAINTS = {
    "qualitative": {
        "min_words": 5,
        "reject_numeric_only": True,
        "forbidden_patterns": ["numeric_only"],
    },
    "quantitative": {
        "must_contain_number": True,
        "must_match_uom": True,
    },
    "boolean": {
        "allowed_values": ["yes", "no", "true", "false", "1", "0"],
    },
}


# =============================================================================
# QUESTION RESOURCE MAP
# Every subcategory × attribute combination maps to:
#   preferred_sources, keywords, table_sections, narrative_sections,
#   answer_mode, validation_rules
# =============================================================================

RESOURCE_MAP: Dict[str, Dict[str, Dict[str, Any]]] = {

    # ── ENVIRONMENT ──────────────────────────────────────────────────────────

    "energy": {
        "policy": {
            "preferred_sources": [
                "BRSR Section B Principle 6",
                "Sustainability Report Energy Policy section",
                "Energy Policy PDF",
                "ESG Governance Section",
                "ISO 50001 certification",
            ],
            "keywords": ["energy policy", "energy management policy",
                         "ISO 50001", "energy governance", "energy efficiency policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "GRI 302"],
        },
        "actions": {
            "preferred_sources": [
                "Sustainability initiatives section",
                "Climate strategy / decarbonization roadmap",
                "Operational excellence section",
                "Energy efficiency projects",
                "CDP Climate Change C8",
            ],
            "keywords": ["waste heat recovery", "PAT scheme", "ISO 50001",
                         "energy audit", "renewable electricity", "energy optimization",
                         "smart metering", "energy efficiency", "LED lighting",
                         "HVAC optimization", "variable frequency drive"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "CDP C8", "GRI 302"],
        },
        "performance": {
            "preferred_sources": [
                "BRSR Annexure / KPI tables",
                "GRI 302 Energy data tables",
                "Environmental Performance Table",
                "Energy Consumption Table",
                "CDP Climate Change C8",
            ],
            "keywords": ["GJ", "energy consumed", "fuel consumption",
                         "electricity consumption", "total energy", "MWh", "TJ"],
            "table_sections": ["Energy", "Resource Consumption",
                                "Environmental KPIs", "GRI 302"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 302", "CDP C8", "Ecovadis"],
        },
        "strategic": {
            "preferred_sources": [
                "Annual report energy audit section",
                "Sustainability report assurance statement",
                "Renewable energy arrangement section",
            ],
            "keywords": ["energy audit", "assurance", "renewable arrangement",
                         "captive plant", "PPA", "REC", "green power"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "GRI 302"],
        },
    },

    "water": {
        "policy": {
            "preferred_sources": [
                "Water stewardship section",
                "Water management policy",
                "Water conservation policy",
                "BRSR Section B P6",
                "CDP Water Security W1",
            ],
            "keywords": ["water policy", "water management", "water stewardship",
                         "responsible water use", "water conservation policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "CDP Water W1", "GRI 303"],
        },
        "actions": {
            "preferred_sources": [
                "Water stewardship section",
                "Water conservation initiatives",
                "Operational sustainability section",
                "Wastewater management section",
                "CDP Water Security W4",
            ],
            "keywords": ["rainwater harvesting", "zero liquid discharge", "ZLD",
                         "water recycling", "wastewater treatment", "closed loop cooling",
                         "reuse of treated water", "water efficiency", "water reuse",
                         "water risk assessment", "leak detection"],
            "answer_mode": "semantic_extraction",
            "forbidden_patterns": ["numeric_only"],
            "frameworks": ["BRSR", "Ecovadis", "CDP Water W4", "GRI 303"],
        },
        "performance": {
            "preferred_sources": [
                "Water accounting tables",
                "BRSR P6 Essential Indicators Q3 Q4",
                "GRI 303 Water and Effluents",
                "CDP Water Security W8",
                "Environmental KPI table",
            ],
            "keywords": ["water withdrawal", "KL", "ML", "m3", "water discharged",
                         "water consumption", "water intensity", "effluent"],
            "table_sections": ["Water", "Water Management", "GRI 303",
                                "Environmental KPIs", "Resource Consumption"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 303", "CDP Water W8", "Ecovadis"],
        },
        "strategic": {
            "preferred_sources": [
                "Water assessment study section",
                "Rainwater harvesting section",
                "Water audit section",
                "CDP Water Security W2",
            ],
            "keywords": ["water assessment", "rainwater harvesting", "water audit",
                         "wastewater testing", "water stressed", "water risk"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "CDP Water", "GRI 303"],
        },
    },

    "waste": {
        "policy": {
            "preferred_sources": [
                "Waste management policy",
                "BRSR Section B P6",
                "Circular economy policy",
                "Ecovadis waste section",
            ],
            "keywords": ["waste policy", "waste management policy",
                         "circular economy policy", "hazardous waste policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "GRI 306"],
        },
        "actions": {
            "preferred_sources": [
                "Circular economy section",
                "Waste reduction initiatives",
                "Hazardous waste management section",
                "Plastic recovery section",
                "CDP Waste module",
            ],
            "keywords": ["circular economy", "recycling", "hazardous waste management",
                         "plastic recovery", "waste minimization", "co-processing",
                         "landfill diversion", "waste segregation", "material reuse",
                         "waste-to-energy", "composting"],
            "answer_mode": "semantic_extraction",
            "forbidden_patterns": ["numeric_only"],
            "frameworks": ["BRSR", "Ecovadis", "CDP Waste", "GRI 306"],
        },
        "performance": {
            "preferred_sources": [
                "Waste data tables",
                "BRSR P6 Essential Indicators Q9",
                "GRI 306 Waste",
                "Environmental KPI table",
                "CDP Waste module",
            ],
            "keywords": ["waste generated", "waste disposed", "hazardous waste",
                         "landfill", "recycling rate", "tonnes", "MT",
                         "plastic waste", "e-waste", "incineration"],
            "table_sections": ["Waste", "Waste Management", "GRI 306",
                                "Environmental KPIs"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 306", "CDP Waste", "Ecovadis"],
        },
        "strategic": {
            "preferred_sources": [
                "Waste audit section",
                "Waste vendor audit section",
                "BRSR P6 Q9 assurance note",
            ],
            "keywords": ["waste audit", "waste assurance", "vendor audit",
                         "recycler audit", "waste certification"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
    },

    "ghg": {
        "policy": {
            "preferred_sources": [
                "GHG / Climate policy section",
                "BRSR Section B P6",
                "CDP Climate Change C1",
                "Ecovadis GHG section",
            ],
            "keywords": ["GHG policy", "climate policy", "emissions management policy",
                         "carbon management", "decarbonization policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "CDP C1", "GRI 305"],
        },
        "actions": {
            "preferred_sources": [
                "Net zero roadmap",
                "Decarbonization strategy",
                "Renewable energy transition section",
                "CDP Climate Change C4 C12",
                "SBTi commitment page",
                "S&P CSA climate strategy",
            ],
            "keywords": ["net zero", "decarbonization", "renewable transition",
                         "emission reduction projects", "green hydrogen", "CCUS",
                         "low carbon", "carbon offset", "SBTi", "science based target",
                         "carbon neutral", "climate action"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "CDP C4 C12", "GRI 305", "Ecovadis", "S&P CSA"],
        },
        "performance": {
            "preferred_sources": [
                "GHG emissions data table",
                "BRSR P6 Essential Indicators Q7",
                "GRI 305 Emissions",
                "CDP Climate Change C6 C7",
                "S&P CSA GHG intensity table",
                "Environmental KPI table",
            ],
            "keywords": ["Scope 1", "Scope 2", "Scope 3", "tCO2e", "tonne CO2",
                         "GHG emissions", "carbon emissions", "emission intensity"],
            "table_sections": ["GHG Emissions", "Scope 1", "Scope 2", "Scope 3",
                                "GRI 305", "Environmental KPIs", "CDP C6"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 305", "CDP C6 C7", "Ecovadis", "S&P CSA"],
        },
        "strategic": {
            "preferred_sources": [
                "Scope 3 methodology section",
                "GHG reporting practices section",
                "CDP Climate Change C5",
            ],
            "keywords": ["Scope 3 method", "average data method", "spend-based",
                         "GHG protocol", "third party verification", "GHG inventory"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "CDP C5", "GRI 305", "Ecovadis"],
        },
    },

    "air pollution": {
        "policy": {
            "preferred_sources": [
                "Air quality policy section",
                "BRSR Section B P6",
                "Environmental compliance section",
            ],
            "keywords": ["air pollution policy", "air quality management",
                         "emission control policy", "stack emission policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "GRI 305"],
        },
        "actions": {
            "preferred_sources": [
                "Air pollution control section",
                "BRSR P6 Leadership Indicators Q4",
                "Environmental monitoring section",
                "CPCB SPCB compliance section",
            ],
            "keywords": ["scrubbers", "ESP", "bag filter", "CEMS", "stack monitoring",
                         "air pollution control", "EIA", "emission reduction",
                         "preventive maintenance", "emergency response plan"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "GRI 305", "CPCB"],
        },
        "performance": {
            "preferred_sources": [
                "Air emissions data table",
                "BRSR P6 Essential Indicators Q6",
                "GRI 305 air quality",
                "CPCB SPCB monitoring reports",
            ],
            "keywords": ["NOx", "SOx", "PM", "particulate matter", "mg/Nm3",
                         "ppm", "air emissions", "stack emissions"],
            "table_sections": ["Air Emissions", "GRI 305", "Environmental KPIs",
                                "BRSR P6 EI-6"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 305", "Ecovadis", "CPCB"],
        },
        "strategic": {
            "preferred_sources": [
                "Air emissions audit section",
                "BRSR P6 EI-6 assurance note",
            ],
            "keywords": ["air emissions audit", "stack audit", "CEMS assurance"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
    },

    # ── SOCIAL ───────────────────────────────────────────────────────────────

    "employee health and safety": {
        "policy": {
            "preferred_sources": [
                "Occupational Health and Safety policy",
                "EHS policy document",
                "BRSR Section B P3",
                "Safety governance section",
            ],
            "keywords": ["OHS policy", "EHS policy", "safety policy",
                         "occupational health", "ISO 45001", "safety governance"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "OSHA", "GRI 403"],
        },
        "actions": {
            "preferred_sources": [
                "Safety training section",
                "Health and safety initiatives",
                "BRSR P3 Essential Indicators Q10",
                "OSHA compliance section",
            ],
            "keywords": ["safety training", "PPE", "incident prevention",
                         "toolbox talks", "safety audits", "behavioral safety",
                         "health check-up", "risk assessment", "emergency plan"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "OSHA", "GRI 403"],
        },
        "performance": {
            "preferred_sources": [
                "Safety KPI table",
                "BRSR P3 Essential Indicators Q11",
                "GRI 403 Occupational Health and Safety",
                "Annual safety report",
            ],
            "keywords": ["LTIFR", "fatalities", "working hours", "TRIR",
                         "lost time injuries", "LTI", "near miss", "first aid"],
            "table_sections": ["Safety KPIs", "GRI 403", "BRSR P3 EI-11",
                                "Health and Safety Data"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 403", "OSHA", "Ecovadis"],
        },
    },

    "employee welfare": {
        "policy": {
            "preferred_sources": [
                "Working conditions policy",
                "Employee welfare policy",
                "BRSR Section B P3",
            ],
            "keywords": ["working conditions policy", "employee welfare",
                         "HR policy", "people policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
        "actions": {
            "preferred_sources": [
                "Employee benefits section",
                "People and culture section",
                "BRSR P3 Essential Indicators Q15",
            ],
            "keywords": ["flexible working", "parental leave", "health insurance",
                         "employee satisfaction", "bonus scheme", "remote work",
                         "family friendly", "childcare"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis"],
        },
        "performance": {
            "preferred_sources": [
                "Employee benefits data table",
                "BRSR P3 Essential Indicators Q1",
                "HR KPI table",
            ],
            "keywords": ["health insurance", "accident insurance", "maternity",
                         "paternity", "retirement benefits", "well-being cost", "INR"],
            "table_sections": ["Employee Benefits", "BRSR P3 EI-1"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR"],
        },
        "strategic": {
            "preferred_sources": [
                "Employee benefits section",
                "BRSR P3 Essential Indicators Q1 Q2",
            ],
            "keywords": ["health insurance", "accident insurance", "maternity",
                         "paternity", "paid time off", "flexible hours",
                         "retirement", "gym", "100% employees"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR"],
        },
    },

    "human rights": {
        "policy": {
            "preferred_sources": [
                "Human rights policy",
                "Anti child labour policy",
                "Forced labour policy",
                "BRSR Section B P5",
                "Ecovadis human rights section",
            ],
            "keywords": ["human rights policy", "anti child labour",
                         "forced labour policy", "ethical sourcing",
                         "modern slavery", "human trafficking policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis", "GRI 408 409"],
        },
        "actions": {
            "preferred_sources": [
                "Supplier due diligence section",
                "Human rights assessment section",
                "BRSR P5 Essential Indicators Q10",
                "Grievance mechanism section",
            ],
            "keywords": ["supplier due diligence", "human rights assessment",
                         "grievance mechanism", "worker awareness",
                         "age verification", "remediation procedure"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "GRI 408 409"],
        },
    },

    "diversity and equal opportunity": {
        "actions": {
            "preferred_sources": [
                "DEI initiatives section",
                "Diversity and inclusion section",
                "BRSR Section A Q21",
                "Ecovadis diversity section",
            ],
            "keywords": ["diversity", "equity", "inclusion", "DEI",
                         "gender equality", "disability inclusion",
                         "anti-discrimination", "harassment prevention"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "GRI 405"],
        },
        "performance": {
            "preferred_sources": [
                "Diversity KPI table",
                "BRSR P5 Essential Indicators Q2 Q7",
                "GRI 405 Diversity",
                "Wage data table",
            ],
            "keywords": ["wages", "INR", "POSH complaints", "women percentage",
                         "gender diversity", "pay equity"],
            "table_sections": ["Diversity KPIs", "GRI 405", "BRSR P5 EI-2"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "GRI 405"],
        },
        "policy": {
            "preferred_sources": [
                "DEI policy",
                "BRSR Section B P3 P5",
            ],
            "keywords": ["diversity policy", "inclusion policy", "DEI policy",
                         "equal opportunity policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
        "strategic": {
            "preferred_sources": [
                "Wage audit section",
                "BRSR P5 assurance note",
            ],
            "keywords": ["wage audit", "assurance provider", "external agency"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR"],
        },
    },

    "sustainable procurement": {
        "policy": {
            "preferred_sources": [
                "Sustainable procurement policy",
                "BRSR Section B P5 P6",
                "Ecovadis procurement section",
            ],
            "keywords": ["sustainable procurement policy", "responsible sourcing",
                         "supplier sustainability policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
        "actions": {
            "preferred_sources": [
                "Supplier ESG assessment section",
                "Supply chain sustainability section",
                "BRSR P2 Leadership Indicators Q2",
                "Ecovadis supplier section",
            ],
            "keywords": ["supplier ESG assessment", "supplier code of conduct",
                         "vendor audit", "supplier sustainability training",
                         "supplier diversity", "sustainable procurement"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis", "GRI 308 414"],
        },
        "performance": {
            "preferred_sources": [
                "Procurement KPI table",
                "BRSR P8 Essential Indicators Q4",
                "Ecovadis supplier data",
            ],
            "keywords": ["MSME procurement", "local sourcing", "supplier code",
                         "supplier audit coverage", "INR", "percentage"],
            "table_sections": ["Procurement KPIs", "BRSR P8 EI-4"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "Ecovadis"],
        },
    },

    "inclusive development": {
        "performance": {
            "preferred_sources": [
                "Inclusive development KPI table",
                "BRSR P8 Essential Indicators Q5",
                "CSR report",
            ],
            "keywords": ["rural wages", "semi-urban wages", "urban wages",
                         "metropolitan wages", "INR", "location-wise wages"],
            "table_sections": ["BRSR P8 EI-5", "Inclusive Development"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR"],
        },
    },

    "social dialogue": {
        "policy": {
            "preferred_sources": [
                "Social dialogue policy",
                "BRSR Section B P3",
                "Industrial relations section",
            ],
            "keywords": ["social dialogue policy", "collective bargaining",
                         "worker consultation", "union policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
    },

    "employee training and development": {
        "policy": {
            "preferred_sources": [
                "Training and development policy",
                "BRSR Section B P3",
                "Learning and development section",
            ],
            "keywords": ["training policy", "career development policy",
                         "L&D policy", "skill development policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["BRSR", "Ecovadis"],
        },
        "actions": {
            "preferred_sources": [
                "Training initiatives section",
                "BRSR P3 Essential Indicators Q8",
                "Learning and development section",
            ],
            "keywords": ["skill development", "performance assessment",
                         "career development", "internal mobility",
                         "training programs", "upskilling"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["BRSR", "Ecovadis"],
        },
    },

    # ── GOVERNANCE ───────────────────────────────────────────────────────────

    "ethics": {
        "policy": {
            "preferred_sources": [
                "Anti-bribery policy",
                "Code of conduct",
                "Ethics policy",
                "Information security policy",
                "Ecovadis ethics section",
                "DJSI ethics section",
            ],
            "keywords": ["anti-bribery policy", "conflict of interest policy",
                         "fraud policy", "money laundering policy",
                         "information security policy", "code of conduct",
                         "ethics policy", "anti-corruption policy"],
            "answer_mode": "boolean_detection",
            "frameworks": ["Ecovadis", "DJSI", "S&P CSA", "GRI 205"],
        },
        "actions": {
            "preferred_sources": [
                "Anti-corruption initiatives section",
                "Whistleblower mechanism section",
                "Information security section",
                "Ecovadis ethics section",
                "S&P CSA anti-corruption",
            ],
            "keywords": ["whistleblower", "ethics training", "compliance certification",
                         "anti-corruption audit", "corruption risk assessment",
                         "information security training", "incident response plan"],
            "answer_mode": "semantic_extraction",
            "frameworks": ["Ecovadis", "DJSI", "S&P CSA", "GRI 205"],
        },
        "performance": {
            "preferred_sources": [
                "Ethics KPI table",
                "BRSR P1 P9 Essential Indicators",
                "Ecovadis ethics data",
                "S&P CSA ethics data",
            ],
            "keywords": ["ethics training percentage", "whistleblower reports",
                         "corruption incidents", "data breaches", "accounts payable",
                         "procurement cost", "INR", "number"],
            "table_sections": ["Ethics KPIs", "BRSR P1 EI-8 EI-9",
                                "BRSR P9 EI-7", "Governance Data"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "Ecovadis", "S&P CSA"],
        },
    },

    "fair trade practices": {
        "performance": {
            "preferred_sources": [
                "Procurement and sales data table",
                "BRSR P1 Essential Indicators Q8 Q9",
                "Financial statements",
                "Annual report financial highlights",
            ],
            "keywords": ["total purchases", "trading houses", "dealers",
                         "distributors", "related parties", "loans advances",
                         "investments", "INR", "total sales"],
            "table_sections": ["BRSR P1 EI-8 EI-9", "Procurement Data",
                                "Sales Data", "Financial KPIs"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR"],
        },
    },

    "disclosure": {
        "strategic": {
            "preferred_sources": [
                "Sustainability reporting section",
                "Assurance statement",
                "DJSI / S&P CSA disclosure section",
                "Annual report governance section",
            ],
            "keywords": ["BRSR", "GRI", "SASB", "TCFD", "sustainability disclosure",
                         "external assurance", "third party assurance"],
            "answer_mode": "boolean_detection",
            "frameworks": ["DJSI", "S&P CSA", "GRI 2"],
        },
    },

    "board diversity": {
        "policy": {
            "preferred_sources": [
                "Board diversity policy",
                "Board independence statement",
                "DJSI / S&P CSA governance section",
                "Corporate governance report",
            ],
            "keywords": ["board diversity policy", "independence statement",
                         "board composition policy", "director independence"],
            "answer_mode": "boolean_detection",
            "frameworks": ["DJSI", "S&P CSA", "GRI 2"],
        },
        "performance": {
            "preferred_sources": [
                "Board composition table",
                "BRSR Section A Q21",
                "Corporate governance report",
                "DJSI board data",
            ],
            "keywords": ["women on board", "female directors", "board diversity",
                         "independent directors", "percentage women", "CEO compensation"],
            "table_sections": ["Board Composition", "BRSR Section A Q21",
                                "Corporate Governance Data"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "DJSI", "S&P CSA", "GRI 2"],
        },
    },

    "general": {
        "performance": {
            "preferred_sources": [
                "Executive compensation section",
                "BRSR Section A",
                "Annual report financial highlights",
                "DJSI compensation data",
            ],
            "keywords": ["CEO compensation", "median compensation",
                         "pay ratio", "executive pay", "INR"],
            "table_sections": ["Compensation Data", "BRSR Section A"],
            "answer_mode": "table_lookup",
            "frameworks": ["BRSR", "DJSI", "S&P CSA"],
        },
        "strategic": {
            "preferred_sources": [
                "ESG governance section",
                "Board oversight section",
                "DJSI governance section",
            ],
            "keywords": ["ESG oversight", "board level ESG", "sustainability committee",
                         "ESG governance", "executive ESG responsibility"],
            "answer_mode": "boolean_detection",
            "frameworks": ["DJSI", "S&P CSA"],
        },
    },

    "materiality": {
        "actions": {
            "preferred_sources": [
                "Materiality assessment section",
                "DJSI materiality section",
                "Sustainability report materiality matrix",
            ],
            "keywords": ["materiality assessment", "material issues",
                         "double materiality", "stakeholder engagement",
                         "material topics"],
            "answer_mode": "boolean_detection",
            "frameworks": ["DJSI", "S&P CSA", "GRI 3"],
        },
    },

    "risk assessment": {
        "actions": {
            "preferred_sources": [
                "Risk management section",
                "Enterprise risk management",
                "DJSI risk section",
                "Annual report risk section",
            ],
            "keywords": ["risk management", "risk culture", "ERM",
                         "risk strategy", "risk governance"],
            "answer_mode": "boolean_detection",
            "frameworks": ["DJSI", "S&P CSA", "GRI 2"],
        },
    },
}

# Fallback resource map for subcategories not explicitly listed
DEFAULT_RESOURCE = {
    "preferred_sources": [
        "BRSR annual report BSE/NSE",
        "GRI sustainability report",
        "CDP questionnaire response",
        "Ecovadis scorecard",
        "DJSI S&P CSA response",
        "Company sustainability report",
        "Annual report",
    ],
    "keywords": [],
    "table_sections": [],
    "answer_mode": "semantic_extraction",
    "frameworks": ["BRSR", "GRI", "CDP", "Ecovadis", "DJSI", "S&P CSA"],
}


def get_resource(subcategory: str, attribute: str) -> Dict[str, Any]:
    """Look up the resource map for a given subcategory + attribute."""
    sub_key  = subcategory.lower().strip()
    attr_key = attribute.lower().strip()
    sub_map  = RESOURCE_MAP.get(sub_key, {})
    return sub_map.get(attr_key, sub_map.get("performance", DEFAULT_RESOURCE))


# =============================================================================
# XLSX LOADER
# =============================================================================

def _clean(raw: Any) -> str:
    if not raw:
        return ""
    return re.sub(r"\s{2,}", " ", re.sub(r"\n+", " | ", str(raw).strip()))


def _question_type(attribute: str, uom: str) -> str:
    attr = attribute.lower().strip()
    uom_l = uom.lower().strip()
    if attr in QUANTITATIVE_ATTRS or (uom_l and uom_l not in {"-", "na", "n/a", ""}):
        return "quantitative"
    if attr in QUALITATIVE_ATTRS:
        return "qualitative"
    return "boolean"


def load_framework_questions(xlsx_path: Path) -> List[Dict[str, Any]]:
    print(f"[xlsx] Loading {xlsx_path.name} ...")
    wb = openpyxl.load_workbook(str(xlsx_path))
    qs: List[Dict[str, Any]] = []

    def _add(ws, p_col, sub_col, attr_col, q_col, uom_col,
             fw_col, opt_col, scheme_col, max_col, min_col):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0] or not row[q_col]:
                continue
            pillar    = str(row[p_col]    or "").strip()
            subcateg  = str(row[sub_col]  or "").strip()
            attribute = str(row[attr_col] or "").strip()
            uom       = str(row[uom_col]  or "").strip()
            q_type    = _question_type(attribute, uom)
            resource  = get_resource(subcateg, attribute)
            qs.append({
                "sno":              row[0],
                "esg_pillar":       pillar,
                "esg_category":     ESG_CATEGORY_LABELS.get(pillar, pillar),
                "esg_subcategory":  subcateg,
                "attribute":        attribute,
                "question_type":    q_type,
                "question":         str(row[q_col]).strip(),
                "uom":              uom,
                "framework_refs":   [f.strip() for f in
                                     str(row[fw_col] or "").split(",") if f.strip()],
                "response_options": _clean(row[opt_col]),
                "scoring_scheme":   _clean(row[scheme_col]),
                "max_score":        float(row[max_col]) if row[max_col] else 0.0,
                "min_score":        float(row[min_col]) if row[min_col] else 0.0,
                # Resource map fields
                "answer_mode":      resource.get("answer_mode", "semantic_extraction"),
                "preferred_sources": resource.get("preferred_sources", []),
                "search_keywords":  resource.get("keywords", []),
                "table_sections":   resource.get("table_sections", []),
                "target_frameworks": resource.get("frameworks", []),
                "forbidden_patterns": resource.get("forbidden_patterns", []),
                "semantic_constraints": SEMANTIC_CONSTRAINTS.get(q_type, {}),
            })

    if "Master questionnaire - Advanced" in wb.sheetnames:
        _add(wb["Master questionnaire - Advanced"],
             p_col=1, sub_col=2, attr_col=3, q_col=4, uom_col=5,
             fw_col=8, opt_col=12, scheme_col=15, max_col=16, min_col=17)
    if "Sheet8" in wb.sheetnames:
        _add(wb["Sheet8"],
             p_col=1, sub_col=2, attr_col=3, q_col=4, uom_col=5,
             fw_col=8, opt_col=12, scheme_col=15, max_col=16, min_col=17)
    if "Sheet9" in wb.sheetnames:
        _add(wb["Sheet9"],
             p_col=1, sub_col=2, attr_col=3, q_col=4, uom_col=5,
             fw_col=8, opt_col=12, scheme_col=14, max_col=15, min_col=15)

    print(f"[xlsx] {len(qs)} questions loaded")
    return qs


# =============================================================================
# LLM CLIENT
# =============================================================================

class LLMClient:
    def __init__(self):
        self._llm = None

    def _init(self):
        if self._llm is None:
            from backend.services.vertex_ai_mime import VertexAILLM
            self._llm = VertexAILLM()

    def call(self, prompt: str, model: str, grounded: bool = False) -> str:
        self._init()
        try:
            return (self._llm.query(
                prompt=prompt, model_name=model, enable_grounding=grounded
            ) or "").strip()
        except Exception as exc:
            print(f"    [llm:{model}] {exc}")
            return ""

    @staticmethod
    def extract_json(raw: str) -> Optional[Dict[str, Any]]:
        if not raw:
            return None
        for marker in ("**Sources", "Sources (live web)", "\n---"):
            if marker in raw:
                raw = raw[:raw.index(marker)]
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            return None
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            return None


# =============================================================================
# STAGE 1 — SCRAPER  (gemini-2.5-flash + Google Search grounding)
# =============================================================================

def scrape_question(
    q: Dict[str, Any],
    company: str,
    llm: LLMClient,
    model: str = "gemini-2.5-flash",
) -> Dict[str, Any]:
    """
    Classified extraction pipeline:
      Question → answer_mode → section routing → source selection → extraction
    """
    mode     = q["answer_mode"]
    sources  = "\n".join(f"  - {s}" for s in q["preferred_sources"]) or "  - Company sustainability report"
    keywords = ", ".join(q["search_keywords"]) if q["search_keywords"] else "ESG data"
    tables   = ", ".join(q["table_sections"])  if q["table_sections"]  else ""
    fw_list  = ", ".join(q["target_frameworks"]) if q["target_frameworks"] else "BRSR, GRI, CDP"
    uom_hint = f" Expected unit: {q['uom']}." if q["uom"] and q["uom"] not in {"-","NA",""} else ""
    constraints = q.get("semantic_constraints", {})

    # Mode-specific extraction instructions
    if mode == "table_lookup":
        extraction_instruction = (
            f"EXTRACTION MODE: table_lookup\n"
            f"Look specifically in these table sections: {tables or 'Environmental/Social/Governance KPI tables'}\n"
            f"Extract the EXACT numeric value with unit.{uom_hint}\n"
            f"If two years of data exist, report both (current FY and previous FY).\n"
            f"REJECT any non-numeric answer for this question."
        )
    elif mode == "boolean_detection":
        extraction_instruction = (
            f"EXTRACTION MODE: boolean_detection\n"
            f"Determine if the company has/does this: Yes or No.\n"
            f"If Yes, briefly describe the specific policy/certification/mechanism found.\n"
            f"Do NOT return a bare number."
        )
    else:  # semantic_extraction
        extraction_instruction = (
            f"EXTRACTION MODE: semantic_extraction\n"
            f"Extract a DESCRIPTIVE answer — specific initiatives, programs, actions, policies.\n"
            f"Minimum {constraints.get('min_words', 5)} words required.\n"
            f"FORBIDDEN: returning only a number. This question requires qualitative text."
        )

    prompt = f"""You are an ESG data extraction specialist. Find the answer for {company}.

ESG CATEGORY  : {q['esg_category']}
SUBCATEGORY   : {q['esg_subcategory']}
ATTRIBUTE     : {q['attribute']}
QUESTION TYPE : {q['question_type']}
QUESTION      : {q['question']}

SEARCH IN THESE SOURCES (check ALL of them):
{sources}

FRAMEWORKS TO CHECK: {fw_list}
SEARCH KEYWORDS: {keywords}

{extraction_instruction}

SCORING SCHEME: {q['scoring_scheme'] or 'Yes=max, No=0, data present=score, absent=0'}
RESPONSE OPTIONS: {q['response_options'] or 'Yes/No or numeric value'}
MAX SCORE: {q['max_score']}

Apply the scoring scheme to assign achieved_score.

Respond ONLY in this JSON (no markdown, no code fences):
{{
  "scraped_answer": "<exact answer>",
  "answer_sources": ["<url1>", "<url2>"],
  "achieved_score": <0 to {q['max_score']}>,
  "score_rationale": "<one sentence>"
}}

If not publicly disclosed after searching all sources:
{{
  "scraped_answer": "Not publicly disclosed",
  "answer_sources": [],
  "achieved_score": 0,
  "score_rationale": "No public disclosure found"
}}"""

    raw  = llm.call(prompt, model=model, grounded=True)
    data = LLMClient.extract_json(raw) or {}

    score   = round(min(max(float(data.get("achieved_score") or 0), 0.0), q["max_score"]), 2)
    extra   = re.findall(r"https?://[^\s\)\]\"'>]+", raw)
    sources_out = list(dict.fromkeys((data.get("answer_sources") or []) + extra))[:8]

    return {
        **q,
        "scraped_answer":  data.get("scraped_answer", "Not found"),
        "answer_sources":  sources_out,
        "achieved_score":  score,
        "score_rationale": data.get("score_rationale", ""),
    }


# =============================================================================
# STAGE 2 — VALIDATION AGENT  (gemini-2.5-pro)
# =============================================================================

VALIDATION_PROMPT_TEMPLATE = """You are an ESG data validation and correction engine.
Your task is to validate whether the extracted answer semantically matches
the ESG question, attribute type, and expected answer format.

You MUST detect:
- qualitative vs quantitative mismatch
- invalid units
- misplaced numeric values
- fallback leakage from previous rows
- scoring inconsistencies
- structurally invalid answers

Rules:
1. Questions under attributes: Actions | Policies | Initiatives | Strategy | Governance | Commitment
   MUST contain descriptive textual answers.

2. Questions under attributes: Performance | Consumption | Withdrawal | Emissions | Energy | Waste | Intensity
   MUST contain numeric values.

3. If answer is numeric AND question expects qualitative explanation:
   - mark INVALID, validation_error = "QUALITATIVE_NUMERIC_MISMATCH"

4. If uom exists AND answer is non-numeric:
   - mark INVALID, validation_error = "NUMERIC_EXPECTED"

5. Detect repeated fallback leakage:
   If answer is identical to a previous unrelated question's answer,
   mark suspicious_repetition = true

6. If answer format is invalid:
   - achieved_score must become 0
   - scoring_status = "INVALID_ANSWER_FORMAT"

7. NEVER allow:
   - numeric values for action/policy questions
   - narrative text for quantitative KPI questions

8. Use semantic understanding, not just keyword matching.

---
COMPANY       : {company}
ESG CATEGORY  : {esg_category}
SUBCATEGORY   : {esg_subcategory}
ATTRIBUTE     : {attribute}
QUESTION TYPE : {question_type}
QUESTION      : {question}
UOM           : {uom}
ANSWER MODE   : {answer_mode}
SCORING SCHEME: {scoring_scheme}
MAX SCORE     : {max_score}

SCRAPED ANSWER  : {scraped_answer}
ACHIEVED SCORE  : {achieved_score}
SCORE RATIONALE : {score_rationale}
ANSWER SOURCES  : {answer_sources}

PREVIOUS 5 ANSWERS (for leakage detection):
{prev_answers}
---

Output ONLY this JSON (no markdown):
{{
  "valid": true/false,
  "question_type": "qualitative | quantitative | boolean",
  "validation_error": null or "ERROR_CODE",
  "suspicious_repetition": false,
  "corrected_answer": "<corrected or original answer>",
  "achieved_score": <number>,
  "scoring_status": "VALID | INVALID_ANSWER_FORMAT | SCORE_CORRECTED",
  "status": "confirmed | corrected | flagged",
  "confidence": <0.0 to 1.0>,
  "correction_reason": null or "<explanation>",
  "cross_check_sources": ["<url>"]
}}"""


def validate_question(
    q: Dict[str, Any],
    company: str,
    llm: LLMClient,
    prev_answers: List[str],
    model: str = "gemini-2.5-pro",
) -> Dict[str, Any]:
    prompt = VALIDATION_PROMPT_TEMPLATE.format(
        company=company,
        esg_category=q.get("esg_category", ""),
        esg_subcategory=q.get("esg_subcategory", ""),
        attribute=q.get("attribute", ""),
        question_type=q.get("question_type", ""),
        question=q.get("question", ""),
        uom=q.get("uom", "-"),
        answer_mode=q.get("answer_mode", ""),
        scoring_scheme=q.get("scoring_scheme", ""),
        max_score=q.get("max_score", 0),
        scraped_answer=q.get("scraped_answer", ""),
        achieved_score=q.get("achieved_score", 0),
        score_rationale=q.get("score_rationale", ""),
        answer_sources=json.dumps(q.get("answer_sources", [])),
        prev_answers=json.dumps(prev_answers[-5:]),
    )

    raw  = llm.call(prompt, model=model, grounded=True)
    data = LLMClient.extract_json(raw) or {}

    defaults: Dict[str, Any] = {
        "valid":                True,
        "question_type":        q.get("question_type", ""),
        "validation_error":     None,
        "suspicious_repetition": False,
        "corrected_answer":     q.get("scraped_answer", ""),
        "achieved_score":       q.get("achieved_score", 0),
        "scoring_status":       "VALID",
        "status":               "confirmed",
        "confidence":           0.5,
        "correction_reason":    None,
        "cross_check_sources":  [],
    }
    defaults.update({k: v for k, v in data.items() if v is not None})
    defaults["achieved_score"] = round(
        min(max(float(defaults["achieved_score"]), 0.0), q["max_score"]), 2
    )
    return defaults


# =============================================================================
# SCORING AGGREGATION  — E / S / G final scores + Overall
# =============================================================================

def _esg_profile(score: float) -> str:
    for lo, hi, label in ESG_PROFILES:
        if lo <= score < hi:
            return label
    return "Advanced"


def _pillar_summary(questions: List[Dict[str, Any]]) -> Dict[str, Any]:
    attr: Dict[str, Dict] = {}
    for q in questions:
        a = q.get("attribute", "Other")
        if a not in attr:
            attr[a] = {"achieved": 0.0, "max": 0.0, "count": 0}
        # Use validated score if available, else scraped score
        score = float(
            q.get("validation", {}).get("achieved_score",
            q.get("achieved_score", 0.0)) or 0
        )
        attr[a]["achieved"] += score
        attr[a]["max"]      += float(q.get("max_score", 0.0))
        attr[a]["count"]    += 1

    total_ach = sum(v["achieved"] for v in attr.values())
    total_max = sum(v["max"]      for v in attr.values())

    return {
        "by_attribute": {
            a: {
                "achieved_score": round(v["achieved"], 2),
                "max_score":      round(v["max"], 2),
                "question_count": v["count"],
                "percentage":     round(v["achieved"] / v["max"] * 100, 2)
                                  if v["max"] else 0.0,
            }
            for a, v in attr.items()
        },
        "total_achieved_score": round(total_ach, 2),
        "total_max_score":      round(total_max, 2),
        "pillar_percentage":    round(total_ach / total_max * 100, 2)
                                if total_max else 0.0,
    }


def _compute_overall(pillar_summaries: Dict[str, Dict]) -> Dict[str, Any]:
    """
    ESG score = [(40%xE_norm + 35%xS_norm + 25%xG_norm)
                 / (40%x130 + 35%x148 + 25%x79)] x 100
    """
    num = 0.0
    den = 0.0
    bd  = {}

    for pillar, cfg in PILLAR_WEIGHTS.items():
        s   = pillar_summaries.get(pillar, {})
        ach = s.get("total_achieved_score", 0.0)
        mx  = s.get("total_max_score", 0.0)
        fw  = cfg["framework_max"]
        pw  = cfg["pillar_weight"]

        norm     = (ach / mx * fw) if mx else 0.0
        weighted = norm * pw
        num += weighted
        den += fw * pw

        bd[pillar] = {
            "esg_category":          ESG_CATEGORY_LABELS.get(pillar, pillar),
            "pillar_weight":         pw,
            "framework_max":         fw,
            "raw_achieved":          round(ach, 2),
            "raw_max":               round(mx, 2),
            "raw_percentage":        round(ach / mx * 100, 2) if mx else 0.0,
            "normalized_score":      round(norm, 2),
            "weighted_contribution": round(weighted, 2),
        }

    overall = round((num / den * 100) if den else 0.0, 2)
    return {
        "overall_esg_score": overall,
        "esg_profile":       _esg_profile(overall),
        "pillar_breakdown":  bd,
        "formula": "[(40%xE + 35%xS + 25%xG) / (40%x130 + 35%x148 + 25%x79)] x 100",
    }


def _build_final_scores(pillar_summaries: Dict, overall: Dict) -> Dict[str, Any]:
    """
    Build the clean final_scores block shown at the top of the output.
    E score / S score / G score / Overall ESG score.
    """
    bd = overall["pillar_breakdown"]
    return {
        "environmental_score": {
            "label":            "Environmental (E)",
            "raw_score":        bd.get("Environment", {}).get("raw_achieved", 0),
            "max_raw_score":    bd.get("Environment", {}).get("raw_max", 0),
            "percentage":       bd.get("Environment", {}).get("raw_percentage", 0),
            "normalized_score": bd.get("Environment", {}).get("normalized_score", 0),
            "framework_max":    130,
            "weight":           "40%",
        },
        "social_score": {
            "label":            "Social (S)",
            "raw_score":        bd.get("Social", {}).get("raw_achieved", 0),
            "max_raw_score":    bd.get("Social", {}).get("raw_max", 0),
            "percentage":       bd.get("Social", {}).get("raw_percentage", 0),
            "normalized_score": bd.get("Social", {}).get("normalized_score", 0),
            "framework_max":    148,
            "weight":           "35%",
        },
        "governance_score": {
            "label":            "Governance (G)",
            "raw_score":        bd.get("Governance", {}).get("raw_achieved", 0),
            "max_raw_score":    bd.get("Governance", {}).get("raw_max", 0),
            "percentage":       bd.get("Governance", {}).get("raw_percentage", 0),
            "normalized_score": bd.get("Governance", {}).get("normalized_score", 0),
            "framework_max":    79,
            "weight":           "25%",
        },
        "overall_esg_score": {
            "label":   "Overall ESG Score",
            "score":   overall["overall_esg_score"],
            "out_of":  100,
            "profile": overall["esg_profile"],
            "formula": overall["formula"],
        },
    }


# =============================================================================
# MAIN PIPELINE  — 3 workers scraping, sequential validation
# =============================================================================

def _worker_fn(args: Tuple) -> Tuple[int, Dict[str, Any]]:
    idx, q, company, llm, model = args
    result = scrape_question(q, company, llm, model)
    return idx, result


def process_company(
    company_name: str,
    questions: List[Dict[str, Any]],
    llm: LLMClient,
    scrape_model:   str  = "gemini-2.5-flash",
    validate_model: str  = "gemini-2.5-pro",
    workers:        int  = 3,
    delay:          float = 0.5,
    run_validation: bool = True,
) -> Dict[str, Any]:

    total = len(questions)
    print(f"\n{'='*72}")
    print(f"  Company   : {company_name}")
    print(f"  Questions : {total}  |  Workers: {workers}  |  Validate: {run_validation}")
    print(f"{'='*72}")

    # ── Stage 1: Parallel scraping ────────────────────────────────────────────
    print(f"\n[Stage 1] Scraping with {workers} workers ({scrape_model}) ...")
    scraped: List[Dict[str, Any]] = [{}] * total
    work_args = [(i, q, company_name, llm, scrape_model) for i, q in enumerate(questions)]

    done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_worker_fn, a): a[0] for a in work_args}
        for fut in as_completed(futures):
            idx, res = fut.result()
            scraped[idx] = res
            done += 1
            print(f"  [{done:3d}/{total}] "
                  f"[{res.get('esg_category','?')[:3]}] "
                  f"[{res.get('esg_subcategory','')[:14]:14s}] "
                  f"[{res.get('answer_mode','')[:10]:10s}] "
                  f"{res.get('question','')[:45]}... "
                  f"→ {res.get('achieved_score',0)}/{res.get('max_score',0)}")
            time.sleep(delay)

    # ── Stage 2: Sequential validation ───────────────────────────────────────
    if run_validation:
        print(f"\n[Stage 2] Validating with {validate_model} ...")
        prev_answers: List[str] = []
        for idx, q in enumerate(scraped, 1):
            print(f"  [{idx:3d}/{total}] Validating [{q.get('esg_category','?')[:3]}] "
                  f"{q.get('question','')[:55]}...")
            vr = validate_question(q, company_name, llm, prev_answers, validate_model)
            scraped[idx - 1]["validation"] = vr
            prev_answers.append(str(q.get("scraped_answer", "")))
            if len(prev_answers) > 10:
                prev_answers.pop(0)
            time.sleep(delay)

    # ── Aggregate ─────────────────────────────────────────────────────────────
    pillar_qs: Dict[str, List] = {"Environment": [], "Social": [], "Governance": []}
    all_sources: List[str] = []

    for q in scraped:
        p = q.get("esg_pillar", "")
        if p in pillar_qs:
            pillar_qs[p].append(q)
        for src in q.get("answer_sources", []):
            if src and src not in all_sources:
                all_sources.append(src)

    framework_scoring: Dict[str, Any] = {}
    pillar_summaries:  Dict[str, Any] = {}

    for pillar, qs in pillar_qs.items():
        summary = _pillar_summary(qs)
        pillar_summaries[pillar] = summary
        framework_scoring[pillar] = {
            "esg_category": ESG_CATEGORY_LABELS.get(pillar, pillar),
            "questions":    qs,
            **summary,
        }

    overall      = _compute_overall(pillar_summaries)
    final_scores = _build_final_scores(pillar_summaries, overall)

    # Validation stats
    val_stats = {"confirmed": 0, "corrected": 0, "flagged": 0,
                 "invalid": 0, "total": total}
    for q in scraped:
        v = q.get("validation", {})
        s = v.get("status", "confirmed")
        if s in val_stats:
            val_stats[s] += 1
        if not v.get("valid", True):
            val_stats["invalid"] += 1

    print(f"\n  ✓ E={final_scores['environmental_score']['normalized_score']:.1f}/130  "
          f"S={final_scores['social_score']['normalized_score']:.1f}/148  "
          f"G={final_scores['governance_score']['normalized_score']:.1f}/79  "
          f"Overall={final_scores['overall_esg_score']['score']}/100  "
          f"({final_scores['overall_esg_score']['profile']})")
    if run_validation:
        print(f"  ✓ Validation: confirmed={val_stats['confirmed']}  "
              f"corrected={val_stats['corrected']}  "
              f"flagged={val_stats['flagged']}  "
              f"invalid={val_stats['invalid']}")

    return {
        "company_name":          company_name,
        "timestamp":             datetime.now().isoformat(),
        "scrape_model":          scrape_model,
        "validation_model":      validate_model if run_validation else None,
        "validation_enabled":    run_validation,
        "validation_summary":    val_stats if run_validation else None,
        # ── Final scores (top-level, easy to read) ──
        "final_scores":          final_scores,
        # ── Detailed per-pillar breakdown ──
        "framework_scoring":     framework_scoring,
        "score_breakdown":       overall,
        "data_sources_searched": all_sources[:60],
    }


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="ESG Framework Scraper + Validation Agent",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--company", nargs="+", metavar="NAME", required=True,
                        help='e.g. --company "Tata Steel" "Infosys"')
    parser.add_argument("--output", required=True, metavar="FILE",
                        help="Output JSON path")
    parser.add_argument("--framework-xlsx",
                        default=str(PROJECT_ROOT / "esg scoring framework.xlsx"))
    parser.add_argument("--scrape-model",   default="gemini-2.5-flash")
    parser.add_argument("--validate-model", default="gemini-2.5-pro")
    parser.add_argument("--workers",  type=int,   default=3,
                        help="Parallel scraping workers (default: 3)")
    parser.add_argument("--delay",    type=float, default=0.5,
                        help="Seconds between API calls (default: 0.5)")
    parser.add_argument("--no-validate", action="store_true",
                        help="Skip validation stage (faster)")

    args = parser.parse_args()

    questions = load_framework_questions(Path(args.framework_xlsx))
    llm       = LLMClient()
    results   = []

    for company in args.company:
        res = process_company(
            company_name=company,
            questions=questions,
            llm=llm,
            scrape_model=args.scrape_model,
            validate_model=args.validate_model,
            workers=args.workers,
            delay=args.delay,
            run_validation=not args.no_validate,
        )
        results.append(res)

    if len(results) == 1:
        output_data = results[0]
    else:
        output_data = {
            "metadata": {
                "generated_at":           datetime.now().isoformat(),
                "total_companies":        len(results),
                "total_questions_per_co": len(questions),
                "scrape_model":           args.scrape_model,
                "validation_model":       args.validate_model,
                "workers":                args.workers,
                "frameworks_covered": [
                    "BRSR", "BRSR-C", "BRSR-NC", "GRI", "CDP Climate",
                    "CDP Water", "CDP Forests", "Ecovadis", "DJSI",
                    "S&P CSA", "OSHA", "Greenco",
                ],
                "scoring_formula": (
                    "[(40%xE + 35%xS + 25%xG) / "
                    "(40%x130 + 35%x148 + 25%x79)] x 100"
                ),
                "esg_profiles": {
                    "Foundational": "0-25%",
                    "Developing":   "25-50%",
                    "Established":  "50-75%",
                    "Advanced":     "75-100%",
                },
                "answer_source": (
                    "Google Search grounding via Vertex AI Gemini "
                    "(web-only, no local data)"
                ),
            },
            "companies": results,
        }

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n{'='*72}")
    print(f"  Saved  → {out}")
    print(f"  Companies processed: {len(results)}")
    print(f"{'='*72}\n")


if __name__ == "__main__":
    main()
